import streamlit as st
import pandas as pd
from openpyxl import Workbook
import io
from typing import List, Tuple

# Configuração da página
st.set_page_config(
    page_title="Depro App",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para melhorar a aparência
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: bold;
    }
    .section-header {
        font-size: 1.5rem;
        color: #2e8b57;
        margin-bottom: 1rem;
        border-bottom: 2px solid #2e8b57;
        padding-bottom: 0.5rem;
    }
    .success-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        margin: 1rem 0;
    }
    .warning-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        margin: 1rem 0;
    }
    .info-box {
        padding: 1rem;
        border-radius: 0.5rem;
        background-color: #cce5ff;
        border: 1px solid #99ccff;
        margin: 1rem 0;
    }
    .stButton > button {
        width: 100%;
        background: linear-gradient(90deg, #1f77b4, #2e8b57);
        color: white;
        border: none;
        border-radius: 0.5rem;
        padding: 0.5rem 1rem;
        font-weight: bold;
        transition: all 0.3s ease;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }
    .copyable-text {
        font-family: 'Courier New', monospace;
        background-color: #f8f9fa;
        border: 1px solid #dee2e6;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 0.5rem 0;
        max-height: 400px;
        overflow-y: auto;
        user-select: text;
        cursor: text;
        white-space: pre-wrap;
        line-height: 1.5;
    }
</style>
""", unsafe_allow_html=True)

# Inicialização das variáveis de sessão
if 'somente_fisicos' not in st.session_state:
    st.session_state.somente_fisicos = []
if 'somente_mgi' not in st.session_state:
    st.session_state.somente_mgi = []

# Header principal
st.markdown('<h1 class="main-header">🔧 Depro App - Ferramentas de Processamento</h1>', unsafe_allow_html=True)

# Sidebar para navegação
st.sidebar.title("🎯 Navegação")
opcao = st.sidebar.selectbox(
    "Escolha uma ferramenta:",
    ["🔍 PROCV - Comparador", "➖ Removedor de Zero", "✂️ Extrair Entre Vírgulas", 
     "🗑️ Remover Duplicados", "🔄 Conversor 55→0"]
)

# Função auxiliar para criar botões de download
def criar_download_excel(data, filename, sheet_name="Dados"):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if isinstance(data, dict):
        col = 1
        for header, values in data.items():
            ws.cell(row=1, column=col, value=header)
            for i, value in enumerate(values, start=2):
                ws.cell(row=i, column=col, value=value)
            col += 1
    elif isinstance(data, list):
        for i, value in enumerate(data, start=1):
            ws.cell(row=i, column=1, value=value)
    
    wb.save(output)
    return output.getvalue()

def criar_download_txt(data):
    if isinstance(data, list):
        return "\n".join(map(str, data))
    return str(data)

# ==================== ABA PROCV ====================
if opcao == "🔍 PROCV - Comparador":
    st.markdown('<div class="section-header">🔍 PROCV - Comparador de Listas</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="info-box">
    <strong>ℹ️ Como usar:</strong><br>
    • Cole os comunicadores físicos na primeira caixa<br>
    • Cole os comunicadores do MGI na segunda caixa<br>
    • Clique em "Comparar Listas" para ver as diferenças<br>
    • Use os botões para copiar ou exportar os resultados
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📋 Comunicadores Físicos:**")
        fisicos_text = st.text_area(
            "Cole aqui os comunicadores físicos (um por linha)",
            height=400,
            key="fisicos_input",
            placeholder="12345678901\n12345678902\n..."
        )
    
    with col2:
        st.markdown("**💻 Comunicadores no MGI:**")
        mgi_text = st.text_area(
            "Cole aqui os comunicadores do MGI (um por linha)",
            height=400,
            key="mgi_input",
            placeholder="12345678901\n12345678903\n..."
        )
    
    # Botões de ação
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("🔍 Comparar Listas", type="primary"):
            if fisicos_text.strip() and mgi_text.strip():
                fisicos = set(line.strip() for line in fisicos_text.strip().splitlines() if line.strip())
                mgi = set(line.strip() for line in mgi_text.strip().splitlines() if line.strip())
                
                st.session_state.somente_fisicos = sorted(fisicos - mgi)
                st.session_state.somente_mgi = sorted(mgi - fisicos)
                
                st.success("✅ Comparação realizada com sucesso!")
            else:
                st.error("❌ Por favor, preencha ambas as listas antes de comparar.")
    
    with col2:
        if st.button("🗑️ Limpar Tudo"):
            st.session_state.somente_fisicos = []
            st.session_state.somente_mgi = []
            st.rerun()
    
    with col3:
        if st.session_state.somente_fisicos or st.session_state.somente_mgi:
            excel_data = {
                "Físicos (mas não no MGI)": st.session_state.somente_fisicos + 
                    [""] * max(0, len(st.session_state.somente_mgi) - len(st.session_state.somente_fisicos)),
                "MGI (mas não fisicamente)": st.session_state.somente_mgi + 
                    [""] * max(0, len(st.session_state.somente_fisicos) - len(st.session_state.somente_mgi))
            }
            
            st.download_button(
                label="📊 Exportar Excel",
                data=criar_download_excel(excel_data, "comparacao.xlsx", "Comparação"),
                file_name="comparacao_comunicadores.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    # Exibição dos resultados
    if st.session_state.somente_fisicos or st.session_state.somente_mgi:
        st.markdown("---")
        st.markdown("### 📊 Resultados da Comparação")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**📋 Comunicadores que tem fisicamente, mas não no MGI:**")
            if st.session_state.somente_fisicos:
                resultado_fisicos = "\n".join(st.session_state.somente_fisicos)
                st.markdown("**📋 Resultado (clique e arraste para selecionar/copiar):**")
                st.markdown(f'<div class="copyable-text">{resultado_fisicos}</div>', unsafe_allow_html=True)
                st.download_button(
                    label="📋 Download como TXT",
                    data=resultado_fisicos,
                    file_name="fisicos_nao_mgi.txt",
                    mime="text/plain"
                )
            else:
                st.info("Nenhum comunicador encontrado nesta categoria.")
        
        with col2:
            st.markdown("**💻 Comunicadores que tem no MGI, mas não fisicamente:**")
            if st.session_state.somente_mgi:
                resultado_mgi = "\n".join(st.session_state.somente_mgi)
                st.markdown("**📋 Resultado (clique e arraste para selecionar/copiar):**")
                st.markdown(f'<div class="copyable-text">{resultado_mgi}</div>', unsafe_allow_html=True)
                st.download_button(
                    label="📋 Download como TXT",
                    data=resultado_mgi,
                    file_name="mgi_nao_fisicos.txt",
                    mime="text/plain"
                )
            else:
                st.info("Nenhum comunicador encontrado nesta categoria.")

# ==================== ABA REMOVEDOR ====================
elif opcao == "➖ Removedor de Zero":
    st.markdown('<div class="section-header">➖ Removedor/Adicionador de Zero</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="info-box">
    <strong>ℹ️ Como usar:</strong><br>
    • <strong>Remover Zero:</strong> Remove o primeiro zero de números que começam com 0<br>
    • <strong>Adicionar Zero:</strong> Adiciona um zero no início de números com menos de 12 dígitos<br>
    • Cole os números (um por linha) e escolha a operação desejada
    </div>
    """, unsafe_allow_html=True)
    
    comunicadores_input = st.text_area(
        "📱 Cole abaixo os comunicadores (um por linha):",
        height=400,
        placeholder="01234567890\n01234567891\n...",
        key="removedor_input"
    )
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("➖ Remover Zero", type="primary"):
            if comunicadores_input.strip():
                comunicadores = comunicadores_input.strip().splitlines()
                resultado = [num[1:] if num.startswith('0') else num for num in comunicadores]
                st.session_state.resultado_removedor = "\n".join(resultado)
                st.success("✅ Zeros removidos com sucesso!")
            else:
                st.error("❌ Por favor, insira os comunicadores antes de remover zeros.")
    
    with col2:
        if st.button("➕ Adicionar Zero", type="primary"):
            if comunicadores_input.strip():
                comunicadores = comunicadores_input.strip().splitlines()
                resultado = ['0' + num if len(num) < 12 else num for num in comunicadores]
                st.session_state.resultado_removedor = "\n".join(resultado)
                st.success("✅ Zeros adicionados com sucesso!")
            else:
                st.error("❌ Por favor, insira os comunicadores antes de adicionar zeros.")
    
    with col3:
        if st.button("🗑️ Limpar Tudo"):
            if 'resultado_removedor' in st.session_state:
                del st.session_state.resultado_removedor
            st.rerun()
    
    # Exibição do resultado
    if 'resultado_removedor' in st.session_state:
        st.markdown("---")
        st.markdown("### 📊 Resultado:")
        
        st.markdown("**📋 Resultado (clique e arraste para selecionar/copiar):**")
        st.markdown(f'<div class="copyable-text">{st.session_state.resultado_removedor}</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📋 Download como TXT",
                data=st.session_state.resultado_removedor,
                file_name="comunicadores_processados.txt",
                mime="text/plain"
            )

# ==================== ABA EXTRAIR ====================
elif opcao == "✂️ Extrair Entre Vírgulas":
    st.markdown('<div class="section-header">✂️ Extrair Entre Vírgulas</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="info-box">
    <strong>ℹ️ Como usar:</strong><br>
    • Esta ferramenta extrai o texto que vem após a primeira vírgula<br>
    • Exemplo: "João,12345678901" → "12345678901"<br>
    • Cole os dados (um por linha) e clique em "Extrair"
    </div>
    """, unsafe_allow_html=True)
    
    entrada_extrair = st.text_area(
        "📝 Cole os dados abaixo (um por linha):",
        height=400,
        placeholder="Nome,12345678901\nOutro,12345678902\n...",
        key="extrair_input"
    )
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("✂️ Extrair", type="primary"):
            if entrada_extrair.strip():
                comunicadores = entrada_extrair.strip().splitlines()
                resultado = []
                for comunicador in comunicadores:
                    partes = comunicador.split(',')
                    if len(partes) > 1:
                        resultado.append(partes[1].strip())
                    else:
                        resultado.append("Sem vírgula encontrada")
                
                st.session_state.resultado_extrair = "\n".join(resultado)
                st.success("✅ Extração realizada com sucesso!")
            else:
                st.error("❌ Por favor, insira os dados antes de extrair.")
    
    with col2:
        if st.button("🗑️ Limpar Tudo"):
            if 'resultado_extrair' in st.session_state:
                del st.session_state.resultado_extrair
            st.rerun()
    
    # Exibição do resultado
    if 'resultado_extrair' in st.session_state:
        st.markdown("---")
        st.markdown("### 📊 Resultado:")
        
        st.markdown("**📋 Dados extraídos (clique e arraste para selecionar/copiar):**")
        st.markdown(f'<div class="copyable-text">{st.session_state.resultado_extrair}</div>', unsafe_allow_html=True)
        
        st.download_button(
            label="📋 Download como TXT",
            data=st.session_state.resultado_extrair,
            file_name="dados_extraidos.txt",
            mime="text/plain"
        )

# ==================== ABA DUPLICADOS ====================
elif opcao == "🗑️ Remover Duplicados":
    st.markdown('<div class="section-header">🗑️ Remover Duplicados</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="info-box">
    <strong>ℹ️ Como usar:</strong><br>
    • Cole os números (um por linha) na caixa de entrada<br>
    • Clique em "Remover Duplicados" para processar<br>
    • Os números únicos aparecerão na caixa de resultado
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📥 Entrada - Números com possíveis duplicados:**")
        numeros_input = st.text_area(
            "Insira números, um por linha:",
            height=400,
            placeholder="12345678901\n12345678902\n12345678901\n...",
            key="duplicados_input"
        )
    
    with col2:
        st.markdown("**📤 Resultado - Números únicos:**")
        if 'resultado_duplicados' in st.session_state:
            st.markdown("**📋 Resultado (clique e arraste para selecionar/copiar):**")
            st.markdown(f'<div class="copyable-text">{st.session_state.resultado_duplicados}</div>', unsafe_allow_html=True)
        else:
            st.text_area(
                "Resultado aparecerá aqui...",
                value="",
                height=400,
                disabled=True,
                key="placeholder_duplicados"
            )
    
    # Botões de ação
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("🗑️ Remover Duplicados", type="primary"):
            if numeros_input.strip():
                try:
                    numeros = numeros_input.strip().split('\n')
                    numeros = [int(num.strip()) for num in numeros if num.strip()]
                    resultado = list(set(numeros))
                    resultado.sort()  # Ordenar para melhor visualização
                    st.session_state.resultado_duplicados = "\n".join(map(str, resultado))
                    
                    original_count = len(numeros)
                    unique_count = len(resultado)
                    duplicados_removidos = original_count - unique_count
                    
                    st.success(f"✅ Processamento concluído! Removidos {duplicados_removidos} duplicados.")
                    st.info(f"📊 {original_count} números originais → {unique_count} números únicos")
                except ValueError:
                    st.error("❌ Por favor, insira apenas números válidos.")
            else:
                st.error("❌ Por favor, insira os números antes de remover duplicados.")
    
    with col2:
        if st.button("🗑️ Limpar Tudo"):
            if 'resultado_duplicados' in st.session_state:
                del st.session_state.resultado_duplicados
            st.rerun()
    
    with col3:
        if 'resultado_duplicados' in st.session_state:
            st.download_button(
                label="📋 Download TXT",
                data=st.session_state.resultado_duplicados,
                file_name="numeros_unicos.txt",
                mime="text/plain"
            )

# ==================== ABA CONVERSOR ====================
elif opcao == "🔄 Conversor 55→0":
    st.markdown('<div class="section-header">🔄 Conversor 55→0</div>', unsafe_allow_html=True)
    
    st.markdown("""
    <div class="info-box">
    <strong>ℹ️ Como usar:</strong><br>
    • Esta ferramenta converte números que começam com "55" para o formato "0"<br>
    • Exemplo: "5511987654321" → "011987654321"<br>
    • Remove o "55" e adiciona "0" + DDD + restante do número<br>
    • Cole os números (um por linha) e clique em "Converter"
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**📥 Entrada - Números iniciados com 55:**")
        entrada_convert = st.text_area(
            "Cole aqui os números para conversão:",
            height=450,
            placeholder="5511987654321\n5521987654321\n...",
            key="converter_input"
        )
    
    with col2:
        st.markdown("**📤 Saída - Números convertidos:**")
        if 'resultado_converter' in st.session_state:
            st.markdown("**📋 Resultado (clique e arraste para selecionar/copiar):**")
            st.markdown(f'<div class="copyable-text">{st.session_state.resultado_converter}</div>', unsafe_allow_html=True)
        else:
            st.text_area(
                "Resultado aparecerá aqui...",
                value="",
                height=450,
                disabled=True,
                key="placeholder_converter"
            )
    
    # Botões de ação
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        if st.button("🔄 Converter", type="primary"):
            if entrada_convert.strip():
                entrada_texto = entrada_convert.strip().splitlines()
                saida_texto = []
                validos = 0
                invalidos = 0
                
                for numero in entrada_texto:
                    numero = numero.strip()
                    if numero.startswith("55") and len(numero) > 4:
                        ddd = numero[2:4]
                        restante = numero[4:]
                        convertido = f"0{ddd}{restante}"
                        saida_texto.append(convertido)
                        validos += 1
                    else:
                        saida_texto.append("Inválido")
                        invalidos += 1
                
                st.session_state.resultado_converter = "\n".join(saida_texto)
                st.success(f"✅ Conversão concluída! {validos} válidos, {invalidos} inválidos.")
            else:
                st.error("❌ Por favor, insira os números antes de converter.")
    
    with col2:
        if st.button("🗑️ Limpar Tudo"):
            if 'resultado_converter' in st.session_state:
                del st.session_state.resultado_converter
            st.rerun()
    
    with col3:
        if 'resultado_converter' in st.session_state:
            st.download_button(
                label="📋 Download TXT",
                data=st.session_state.resultado_converter,
                file_name="numeros_convertidos.txt",
                mime="text/plain"
            )
    
    with col4:
        if 'resultado_converter' in st.session_state:
            # Preparar dados para Excel (apenas números válidos)
            dados_excel = [linha for linha in st.session_state.resultado_converter.split('\n') if linha != 'Inválido']
            if dados_excel:
                st.download_button(
                    label="📊 Download Excel",
                    data=criar_download_excel(dados_excel, "convertidos.xlsx", "Números Convertidos"),
                    file_name="numeros_convertidos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# Footer
st.markdown("---")
st.markdown("""
<div style="text-align: center; padding: 2rem; color: #666;">
    <h4>🔧 Depro App</h4>
    <p>Ferramenta desenvolvida para processamento eficiente de dados de comunicadores</p>
    <p><em>Versão Streamlit - Interface moderna e responsiva</em></p>
</div>
""", unsafe_allow_html=True)